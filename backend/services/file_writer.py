from datetime import date, datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook

try:
    from .transformer import TransformResult
except ImportError:
    try:
        from services.transformer import TransformResult
    except ModuleNotFoundError:
        from transformer import TransformResult


MAIN_SHEET_NAME = "ERP Import"

NUMBER_COLUMNS = {
    "Quantity",
    "Unit Price",
    "Amount",
    "Term Amount",
}
DATE_COLUMNS = {"Invoice Date"}


class FileWriter:
    def write(
        self,
        result: TransformResult,
        output_dir: str,
        template_name: str = "erp-output",
    ) -> dict[str, str | None]:
        """
        Writes the output Excel file.
        Returns:
        {
          "output_path": str,
          "output_filename": str,
          "summary_path": str | None
        }
        """
        del template_name

        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%S%M%H")
        output_filename = f"output_{timestamp}.xlsx"
        full_output_path = output_path / output_filename

        workbook = Workbook()
        main_sheet = workbook.active
        main_sheet.title = MAIN_SHEET_NAME

        self._write_main_sheet(main_sheet, result.output_df)

        workbook.save(full_output_path)

        error_report_path = None
        if result.error_count > 0:
            error_report_path = output_path / f"erp_errors_{timestamp}.xlsx"
            self._write_error_report(result.errors, error_report_path)

        return {
            "output_path": str(full_output_path),
            "output_filename": output_filename,
            "summary_path": str(error_report_path) if error_report_path else None,
        }

    def _write_main_sheet(
        self,
        sheet,
        dataframe: pd.DataFrame,
    ) -> None:
        for column_index, column_name in enumerate(dataframe.columns, start=1):
            sheet.cell(row=1, column=column_index, value=column_name)

        for row_index, row in enumerate(dataframe.itertuples(index=False), start=2):
            for column_index, value in enumerate(row, start=1):
                column_name = str(dataframe.columns[column_index - 1])
                sheet.cell(
                    row=row_index,
                    column=column_index,
                    value=_excel_value(value, column_name),
                )


    def _write_error_report(self, errors: list[Any], output_path: Path) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Errors"

        headers = [
            "Row Number",
            "ERP Column",
            "Error Message",
            "Original POS Value",
        ]
        for column_index, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=column_index, value=header)

        for row_index, error in enumerate(errors, start=2):
            parsed_error = _parse_error(error)
            values = [
                parsed_error["row_number"],
                parsed_error["erp_column"],
                parsed_error["error_message"],
                parsed_error["original_pos_value"],
            ]
            for column_index, value in enumerate(values, start=1):
                sheet.cell(row=row_index, column=column_index, value=value)

        workbook.save(output_path)


def _excel_value(value: Any, column_name: str) -> Any:
    if column_name in DATE_COLUMNS:
        return _excel_date_value(value)
    if column_name in NUMBER_COLUMNS:
        return _excel_number_value(value)
    if _is_missing(value):
        return ""
    return value


def _excel_number_value(value: Any) -> int | float | str:
    if _is_missing(value):
        return ""
    if isinstance(value, int | float):
        return value

    text_value = str(value).strip().replace(",", "")
    if text_value == "":
        return ""

    try:
        parsed = float(text_value)
    except ValueError:
        return text_value

    if parsed.is_integer() and str(value).strip().isdigit():
        return int(parsed)
    return parsed


def _excel_date_value(value: Any) -> date | str:
    if _is_missing(value):
        return ""
    if isinstance(value, pd.Timestamp):
        return value.date()
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text_value = str(value).strip()
    for date_format in ("%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text_value, date_format).date()
        except ValueError:
            continue
    return text_value


def _is_missing(value: Any) -> bool:
    if value is None:
        return True
    try:
        return bool(pd.isna(value))
    except (TypeError, ValueError):
        return False


def _parse_error(error: Any) -> dict[str, Any]:
    if isinstance(error, dict):
        return {
            "row_number": error.get("row_number") or error.get("row") or "",
            "erp_column": error.get("erp_column") or error.get("column") or "",
            "error_message": error.get("error_message") or error.get("message") or "",
            "original_pos_value": error.get("original_pos_value")
            or error.get("original_value")
            or "",
        }

    message = str(error)
    row_number = ""
    erp_column = ""

    if message.startswith("Row "):
        row_part, _, rest = message.partition(":")
        row_number = row_part.replace("Row", "").strip()
        message = rest.strip() or message

    if " set to " in message:
        before_default, _, _ = message.partition(" set to ")
        _, _, possible_column = before_default.partition(";")
        erp_column = possible_column.strip()

    return {
        "row_number": row_number,
        "erp_column": erp_column,
        "error_message": str(error),
        "original_pos_value": "",
    }

def _resolve_project_path(relative_path: str) -> Path:
    project_root_path = Path(relative_path)
    if project_root_path.exists():
        return project_root_path
    return Path(__file__).resolve().parents[2] / relative_path


if __name__ == "__main__":
    try:
        from .file_reader import FileReader
        from .transformer import DataTransformer
    except ImportError:
        try:
            from services.file_reader import FileReader
            from services.transformer import DataTransformer
        except ModuleNotFoundError:
            from file_reader import FileReader
            from transformer import DataTransformer

    reader = FileReader()
    read_result = reader.read(str(_resolve_project_path("uploads/input.xlsx")))
    transform_result = DataTransformer().transform(read_result["dataframe"])

    write_result = FileWriter().write(
        transform_result,
        str(_resolve_project_path("outputs")),
    )

    print("output_path:", write_result["output_path"])
    print("output_filename:", write_result["output_filename"])
    print("summary_path:", write_result["summary_path"])
